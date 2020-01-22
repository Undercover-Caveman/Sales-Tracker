#MiCRO

# # # # # # # # # # [PRE-REQUISITES] # # # # # # # # # #

import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def Create_Database():
    table = sqlite3.connect('test.db')
    a = table.cursor()
    a.execute ("create table database (ID, fname, lname, phone, email, contact_method, province, condition, year, make, model, trim, notes, owner, conquest, conquestplus, other, tyear, tmake, tmodel, ttrim, km, lien, value, note2)")
    table.commit()
    table.close()

def Read_Database():
    table = sqlite3.connect('test.db')
    get = table.cursor()
    get.execute("SELECT ID, fname, lname FROM database")
    rows = get.fetchall()
    for row in rows:
        (ID, fname, lname) = tuple(row)
        print (ID, fname, lname)
        
# # # # # # # # # # [CLASSES] # # # # # # # # # #

class Client:
    def __init__(self, fname, lname, phone, email, contact_method, province):
        self.fname = fname
        self.lname = lname
        self.phone = phone
        self.email = email
        self.contact_method = contact_method
        self.province = province

class Car:
    def __init__(self, condition, year, make, model, trim, notes, owner, conquest, conquestplus, other):
        self.condition = condition
        self.year = year
        self.make = make
        self.model = model
        self.trim = trim
        self.notes = notes
        self.owner = owner
        self.conquest = conquest
        self.conquestplus = conquestplus
        self.other = other

class Trade:
    def __init__(self, year, make, model, trim, km, lien, value):
        self.year = year
        self.make = make
        self.model = model
        self.trim = trim
        self.km = km
        self.lien = lien
        self.value = value
                
# # # # # # # # # # [ENGINES] # # # # # # # # # #

def Engine():
    count = 0
    while count == 0:
        alpha = input("""
What would you like to do?
1) Add Customer
2) View All Customers
3) Find Customer
4) Create Worksheet
5) Create Database
6) Quit
""")
        if alpha == ("1"):
            Add_Customer()
        elif alpha == ("2"):
            Read_Database()
        elif alpha == ("3"):
            Find_Customer()
        elif alpha == ("4"):
            Make_Worksheet()
        elif alpha == ("5"):
            Create_Database()
        elif alpha == ('6'):
            print ("Good Bye")
            count += 1
        else:
            print ("Error. Please use #s 1 - 5 ")

def Make_Worksheet():
    alpha=(input(""))
    table = sqlite3.connect('test.db')
    get = table.cursor()
    get.execute ("SELECT * FROM database WHERE ID = {}".format(alpha))
    row = get.fetchone()
    (ID, fname, lname, phone, email, contact_method, province, condition, year, make, model, trim, notes, owner, conquest, conquestplus, other, tyear, tmake, tmodel, ttrim, km, lien, value, note2) = tuple(row) 
    print (ID, fname, lname, phone, email, contact_method, province, condition, year, make, model, trim, notes, owner, conquest, conquestplus, other, tyear, tmake, tmodel, ttrim, km, lien, value, note2)

    Customer = Client(fname, lname, phone, email, contact_method, province )
    Auto = Car(condition, year, make, model, trim, notes, owner, conquest, conquestplus, other)
    Trade_In = Trade(tyear, tmake, tmodel, ttrim, km, lien, value)

    print ("Working")
    wb = load_workbook(filename = 'sheet.xlsx')
    ws = wb.active
    select = PatternFill(start_color='DBD9D8', end_color='DBD9D8', fill_type='solid')

    ws['C8'] = ((Customer.fname) + " " + (Customer.lname))
    ws['C10'] = (Customer.phone)
    ws['G10'] = (Customer.email)

    if Customer.contact_method == ("Walk-in"):
        ws['D12'].fill = (select)
        pass
    elif Customer.contact_method == ("Phone"):
        ws['F12'].fill = (select)
        pass
    elif Customer.contact_method == ("Rapid"):
        ws['H12'].fill = (select)
        pass
    elif Customer.contact_method == ("Service"):
        ws['J12'].fill = (select)
        pass
    else:
        ws['D14'].fill = (select)
        pass

    if Customer.province == ("Ontario"):
        ws['D16'].fill = (select)
        pass
    elif Customer.province == ("Quebec"):
        ws['F16'].fill = (select)
        pass
    else:
        we['H16'].fill = (select)
        pass

    if Auto.condition == ("New"):
        ws['D18'].fill = (select)
        pass
    elif Auto.condition == ("CPO"):
        ws['F18'].fill = (select)
        pass
    elif Auto.condition == ("Used"):
        ws['H18'].fill = select
        pass
    else:
        pass

    ws['E20'] = (Auto.year)
    ws['G20'] = (Auto.make)
    ws['I20'] = (Auto.model)
    ws['K20'] = (Auto.trim)
    ws['E22'] = (Auto.notes)

    if Auto.owner == ("Y"):
        ws['C28'].fill = (select)
        pass
    if Auto.conquest == ("Y"):
        ws['F28'].fill = (select)
        pass
    if Auto.conquestplus == ("Y"):
        ws['I28'].fill = (select)
        pass
    if Auto.other != ("N"):
        ws['E30'].fill = (select)
        ws['E30'] = (Auto.other)
    
    ws['E32'] = (Trade_In.year)
    ws['G32'] = (Trade_In.make)
    ws['I32'] = (Trade_In.model)
    ws['K32'] = (Trade_In.trim)
    ws['E34'] = (Trade_In.km)   
    ws['E36'] = (Trade_In.lien)
    ws['I36'] = (Trade_In.value)
    xray = note2
    ws['A39'] = (xray)

    save = input("Save name? ")
    save = save+(".xlsx")
    wb.save(('%s') % (save))
    print ("done")
                    
def Add_Customer():
    alpha = First_Name()
    bravo = Last_Name()
    charlie = Phone_Number()
    delta = Email()
    echo = Contact_Method()
    foxtrot = Province()

    count = 0
    while count == 0:
        ALPHA = input("Add vehicle of interest? Y/N ")
        if ALPHA.upper() == ("Y"):                 
            golf = Interest_Condition()
            hotel = Interest_Year()
            india = Interest_Make()
            juliette = Interest_Model()
            kilo = Interest_Trim()
            lima = Notes()
            mike = Program_Owner()
            november = Program_Conquest()
            oscar = Program_Conquest_Plus()
            papa = Program_Other()
            count += 1
        elif ALPHA.upper()[0] == ("N"):
            golf = (' ')
            hotel = (' ')
            india = (' ')
            juliette = (' ')
            kilo = (' ')
            lima = (' ')
            mike = (' ')
            november = (' ')
            oscar = (' ')
            papa = (' ')
            count += 1
        else:
            print("Error")
                
    count = 0
    while count == 0:  
        BRAVO = input("Add trade? Y/N ")
        if BRAVO.upper()[0] == ("Y"): 
            quebec = Trade_Year()
            romeo = Trade_Make()
            sierra = Trade_Model()
            tango = Trade_Trim()
            uniform = Trade_KM()
            victor = Trade_Lien()
            whiskey = Trade_Value()
            count += 1
        elif BRAVO.upper()[0] == ("N"):
            quebec = (' ')
            romeo = (' ')
            sierra = (' ')
            tango = (' ')
            uniform = (' ')
            victor = (' ')
            whiskey = (' ')
            count += 1
            pass
        else:
            print("Error")
            
    xray = Notes2()    
    Customer = Client(alpha, bravo, charlie, delta, echo, foxtrot)
    Auto = Car(golf, hotel, india, juliette, kilo, lima, mike, november, oscar, papa)
    Trade_In = Trade(quebec, romeo, sierra, tango, uniform, victor, whiskey)

    table = sqlite3.connect('test.db')
    a = table.cursor()
    a.execute ("\
insert into database (ID, fname, lname, phone, email, contact_method, province, condition, year, make, model, trim, notes, owner, conquest, conquestplus, \
other, tyear, tmake, tmodel, ttrim, km, lien, value, note2) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",(ID(), Customer.fname, \
Customer.lname, Customer.phone, Customer.email, Customer.contact_method, Customer.province, Auto.condition, Auto.year, Auto.make, Auto.model, Auto.trim, \
Auto.notes, Auto.owner, Auto.conquest, Auto.conquestplus, Auto.other, Trade_In.year, Trade_In.make, Trade_In.model, Trade_In.trim, Trade_In.km, \
Trade_In.lien, Trade_In.value, (xray)))
    table.commit()
    table.close()

# # # # # # # # # # [FUNCTIONS] # # # # # # # # # #
        
def First_Name():
    return input("First Name: ")

def Last_Name():
    return input("Last Name: ")

def Phone_Number():
    return input("Phone #: ")

def Email():
    return input("E-Mail: ")

def Contact_Method():
    count = 0
    while count == 0:
        alpha = input ("""
Contach Method:
1)Walk-in
2)Phone
3)Rapid Response
4)Service
5)Other
""")
        if alpha == ("1"):
            count += 1
            return ("Walk-in")
        elif alpha == ("2"):
            count += 1
            return ("Phone")   
        elif alpha == ("3"):
            count += 1
            return ("Rapid")      
        elif alpha == ("4"):
            count += 1
            return ("Service")
        elif alpha == ("5"):
            count += 1
            return input("Method: ")   
        else:
            print ("Error. Please only use #'s 1-5")
        
def Province():
    count = 0
    while count == 0:
        alpha = input ("""
Province of Residence:
1) Ontario
2) Quebec
3) Other
""")
        if alpha == ("1"):
            count += 1
            return ("Ontario")
        elif alpha == ("2"):
            count += 1
            return ("Quebec")
        elif alpha == ("3"):
            count += 1
            return input("Province: ")
        else:
            print ("Please only use #'s 1-3")

def Interest_Condition():
    count = 0
    while count == 0:
        alpha = input ("""
Condition of Vehicle:
1) New
2) CPO
3) Used
""")
        if alpha == ("1"):
            count += 1
            return ("New")
        elif alpha == ("2"):
            count += 1
            return ("CPO")
        elif alpha == ("3"):
            count += 1
            return ("Used")
        else:
            print ("Please only use #'s 1-3")

def Interest_Year():
    return input("Vehicle year: ")

def Interest_Make():
    return input("Vehicle make: ")

def Interest_Model():
    return input("Vehicle model: ")

def Interest_Trim():
    return input("Vehicle trim: ")

def Notes():
    return input("Notes: ")

def Program_Owner():
    count = 0
    while count == 0:
        alpha = input("Volvo owner? Y/N " )
        if alpha.upper()[0] == ("Y"):
            count += 1
            return ("Y")
        elif alpha.upper()[0] == ("N"):
            count += 1
            return ("N")
        else:
            print ("Error. Please return Y or N")
            
def Program_Conquest():
    count = 0
    while count == 0:
        alpha = input("Conquest? Y/N ")
        if alpha.upper()[0] == ("Y"):
            count += 1
            return ("Y")
        elif alpha.upper()[0] == ("N"):
            count += 1
            return ("N")
        else:
            print ("Error. Please return Y or N")

def Program_Conquest_Plus():
    count = 0
    while count == 0:
        alpha = input("Conquest PLUS? Y/N ")
        if alpha.upper()[0] == ("Y"):
            count += 1
            return ("Y")
        elif alpha.upper()[0] == ("N"):
            count += 1
            return ("N")
        else:
            print ("Error. Please return Y or N")

def Program_Other():
    count = 0
    while count == 0:
        alpha = input("Other programs? Y/N ")
        if alpha.upper()[0] == ("Y"):
            count += 1
            return input("Program: ")
        elif alpha.upper()[0] == ("N"):
            count +=1
            return ("N")
        else:
            print ("Error. Please return Y or N")

def Trade_Year():
    return input("Trade year: ")

def Trade_Make():
    return input ("Trade make: ")

def Trade_Model():
    return input ("Trade model: ")

def Trade_Trim():
    return input ("Trade trim: ")

def Trade_KM():
    return input ("Trade KMs: ")

def Trade_Lien():
    return input ("Trade lien: ")

def Trade_Value():
    return input ("Trade value: ")

def Notes2():
    count = (0)
    while count == (0):
        alpha = input("Notes? Y/N ")
        if alpha.upper()[0] == ("Y"):
            return input("Notes: ")
        elif alpha.upper()[0] == ("N"):
            count = 1
            return (" ")
        else:
            print ("Error. Please return Y or N ")
           
def ID():
    table = sqlite3.connect('test.db')
    get = table.cursor()
    get.execute("select max (ID) from database")
    rows = get.fetchall()[0]
    for row in rows:
        alpha = row
        alpha += 1
        return (alpha)

def Find_Customer():
    count = 0
    while count == (0):
        alpha = input("""
How do you want to search?
1) First name
2) Last Name
3) Phone #
4) E-Mail
5) Close
""")
        table = sqlite3.connect('test.db')
        get = table.cursor()
        bravo = input("Search for: ")
        print (bravo)
        if alpha == ("1"):
            get.execute ("SELECT ID, fname, lname, phone, email FROM database WHERE fname='{}'".format (str(bravo)))
            rows = get.fetchall()
            for row in rows:
                (ID, fname, lname, phone, email) = tuple(row)
                print (ID, fname, lname,phone,email)
            count += 1
        elif alpha == ("2"):
            get.execute ("SELECT ID, fname, lname, phone, email FROM database WHERE lname='{}'".format (str(bravo)))
            rows = get.fetchall()
            for row in rows:
                (ID, fname, lname, phone, email) = tuple(row)
                print (ID, fname, lname, phone, email)
            count += 1
        elif alpha == ("3"):
            get.execute ("SELECT ID, fname, lname, phone, email FROM database WHERE phone='{}'".format (str(bravo)))
            rows = get.fetchall()
            for row in rows:
                (ID, fname, lname, phone, email) = tuple(row)
                print (ID, fname, lname, phone, email)
            count += 1
        elif alpha == ("4"):
            get.execute ("SELECT ID, fname, lname, phone, email FROM database WHERE email='{}'".format (str(bravo)))
            rows = get.fetchall()
            for row in rows:
                (ID, fname, lname, phone, email) = tuple(row)
                print (ID, fname, lname, phone, email)
            count += 1
        elif alpha == ("5"):
            quit
            count += 1
        else:
            print ("Pelase only use #'s 1-5")

Engine()
